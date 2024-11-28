import sys

sys.path.append('./')
sys.path.append('../')
sys.path.append('../../')

import matplotlib.pyplot as plt
from openpyxl import load_workbook

# x = list(range(50))


class Line(object):
    data = []
    color = 'r'
    line_style = '-'
    label = 'label'

    def __init__(self, data, color, line_style, label):
        self.data = data
        self.color = color
        self.line_style = line_style
        self.label = label


def draw_line(lines, x_label, y_label, title, file_name):
    for line in lines:
        x = list(range(len(line.data)))
        plt.plot(x, line.data, color=line.color, linestyle=line.line_style, label=line.label)
    plt.xlabel(x_label)
    plt.ylabel(y_label)
    plt.title(title)
    plt.legend()
    plt.savefig('../results/images/' + file_name + '.png', dpi=300)
    plt.close()


def read_excel(file_name, col_idx, color, line_style, label):
    workbook = load_workbook('../results/' + file_name + '.xlsx')
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row[col_idx])
    return Line(data, color, line_style, label)


# local_gray_test_loss = read_excel('local_gray_test_loss_2024_11_21', col_idx=0, color='g', line_style='-', label='local gray test loss')
# local_color_test_loss = read_excel('local_color_test_loss_2024_11_21', col_idx=0, color='b', line_style='-', label='local color test loss')
# local_multi_test_loss = read_excel('local_multi_test_loss_2024_11_21', col_idx=0, color='y', line_style='-', label='local multi test loss')
# mmfl_test_server_loss = read_excel('MMFL_Test_Loss_2024_11_21', col_idx=0, color='r', line_style='-', label='MMFL test server loss')
# mmfl_test_client1_loss = read_excel('MMFL_Test_Loss_2024_11_21', col_idx=1, color='r', line_style='--', label='MMFL test client1 loss')
# mmfl_test_client2_loss = read_excel('MMFL_Test_Loss_2024_11_21', col_idx=2, color='r', line_style='--', label='MMFL test client2 loss')
# mmfl_proto_test_server_loss = read_excel('MMFL_Proto_Test_Loss_2024_11_21', col_idx=0, color='m', line_style='-', label='MMFL proto test server loss')
# mmfl_proto_test_client1_loss = read_excel('MMFL_Proto_Test_Loss_2024_11_21', col_idx=1, color='m', line_style='--', label='MMFL proto test client1 loss')
# mmfl_proto_test_client2_loss = read_excel('MMFL_Proto_Test_Loss_2024_11_21', col_idx=2, color='m', line_style='--', label='MMFL proto test client2 loss')
#
# local_gray_test_acc = read_excel('local_gray_test_acc_2024_11_21', col_idx=0, color='g', line_style='-', label='local gray test acc')
# local_color_test_acc = read_excel('local_color_test_acc_2024_11_21', col_idx=0, color='b', line_style='-', label='local color test acc')
# local_multi_test_acc = read_excel('local_multi_test_acc_2024_11_21', col_idx=0, color='y', line_style='-', label='local multi test acc')
# mmfl_test_server_acc = read_excel('MMFL_Test_ACC_2024_11_21', col_idx=0, color='r', line_style='-', label='MMFL test server acc')
# mmfl_test_client1_acc = read_excel('MMFL_Test_ACC_2024_11_21', col_idx=1, color='r', line_style='--', label='MMFL test client1 acc')
# mmfl_test_client2_acc = read_excel('MMFL_Test_ACC_2024_11_21', col_idx=2, color='r', line_style='--', label='MMFL test client2 acc')
# mmfl_proto_test_server_acc = read_excel('MMFL_Proto_Test_ACC_2024_11_21', col_idx=0, color='m', line_style='-', label='MMFL proto test server acc')
# mmfl_proto_test_client1_acc = read_excel('MMFL_Proto_Test_ACC_2024_11_21', col_idx=1, color='m', line_style='--', label='MMFL proto test client1 acc')
# mmfl_proto_test_client2_acc = read_excel('MMFL_Test_ACC_2024_11_21', col_idx=2, color='m', line_style='--', label='MMFL proto test client2 acc')
#
# loss_lines = [local_gray_test_loss, local_color_test_loss, local_multi_test_loss,
#               mmfl_test_server_loss, mmfl_test_client1_loss, mmfl_test_client2_loss,
#               mmfl_proto_test_server_loss, mmfl_proto_test_client1_loss, mmfl_proto_test_client2_loss]
# acc_lines = [local_gray_test_acc, local_color_test_acc, local_multi_test_acc,
#              mmfl_test_server_acc, mmfl_test_client1_acc, mmfl_test_client2_acc,
#              mmfl_proto_test_server_acc, mmfl_proto_test_client1_acc, mmfl_proto_test_client2_acc]

# mmfl_resnet_test_server_loss = read_excel('MMFL_Test_Loss_2024_11_24', col_idx=0, color='r', line_style='-', label='MMFL resnet test server loss')
# mmfl_resnet_test_client1_loss = read_excel('MMFL_Test_Loss_2024_11_24', col_idx=1, color='r', line_style='--', label='MMFL resnet client1 loss')
# mmfl_resnet_test_client2_loss = read_excel('MMFL_Test_Loss_2024_11_24', col_idx=2, color='r', line_style='--', label='MMFL resnet client2 loss')
# mmfl_resnet_test_server_no_head_loss = read_excel('MMFL_Test_Loss_2024_11_24_no_head', col_idx=0, color='m', line_style='-', label='MMFL resnet test server no head loss')
# mmfl_resnet_test_client1_no_head_loss = read_excel('MMFL_Test_Loss_2024_11_24_no_head', col_idx=1, color='m', line_style='--', label='MMFL resnet test client1 no head loss')
# mmfl_resnet_test_client2_no_head_loss = read_excel('MMFL_Test_Loss_2024_11_24_no_head', col_idx=2, color='m', line_style='--', label='MMFL resnet test client2 no head loss')
#
# mmfl_resnet_test_server_acc = read_excel('MMFL_Test_ACC_2024_11_24', col_idx=0, color='r', line_style='-', label='MMFL resnet test server acc')
# mmfl_resnet_test_client1_acc = read_excel('MMFL_Test_ACC_2024_11_24', col_idx=1, color='r', line_style='--', label='MMFL resnet test client1 acc')
# mmfl_resnet_test_client2_acc = read_excel('MMFL_Test_ACC_2024_11_24', col_idx=2, color='r', line_style='--', label='MMFL resnet test client2 acc')
# mmfl_resnet_test_server_no_head_acc = read_excel('MMFL_Test_ACC_2024_11_24_no_head', col_idx=0, color='m', line_style='-', label='MMFL resnet test server no head acc')
# mmfl_resnet_test_client1_no_head_acc = read_excel('MMFL_Test_ACC_2024_11_24_no_head', col_idx=1, color='m', line_style='--', label='MMFL resnet test client1 no head acc')
# mmfl_resnet_test_client2_no_head_acc = read_excel('MMFL_Test_ACC_2024_11_24_no_head', col_idx=2, color='m', line_style='--', label='MMFL resnet test client2 no head acc')
#
# loss_lines = [mmfl_resnet_test_server_loss, mmfl_resnet_test_client1_loss, mmfl_resnet_test_client2_loss,
#               mmfl_resnet_test_server_no_head_loss, mmfl_resnet_test_client1_no_head_loss, mmfl_resnet_test_client2_no_head_loss]
# acc_lines = [mmfl_resnet_test_server_acc, mmfl_resnet_test_client1_acc, mmfl_resnet_test_client2_acc,
#              mmfl_resnet_test_server_no_head_acc, mmfl_resnet_test_client1_no_head_acc, mmfl_resnet_test_client2_no_head_acc]

# local_simple_resnet_cifar100_multi_loss = read_excel('2024_11_24_local_resnet_cifar100_multi_test_loss', col_idx=0, color='y', line_style='-', label='local simple resnet multi loss')
# local_simple_resnet_cifar100_color_loss = read_excel('2024_11_24_local_resnet_cifar100_color_test_loss', col_idx=0, color='b', line_style='-', label='local simple resnet color loss')
# local_simple_resnet_cifar100_gray_loss = read_excel('2024_11_24_local_resnet_cifar100_gray_test_loss', col_idx=0, color='g', line_style='-', label='local simple resnet gray loss')
# # mmfl_simple_resnet_cifar100_no_client_loss = read_excel('2024_11_25_mmfl_resnet_cifar100_no_client_loss', col_idx=0, color='r', line_style='-', label='MMFL simple resnet no client loss')
# mmfl_simple_resnet_cifar100_loss = read_excel('2024_11_25_mmfl_resnet_cifar100_loss', col_idx=0, color='m', line_style='-', label='MMFL simple resnet loss')
# # mmfl_simple_resnet_cifar100_with_hfl_loss = read_excel('2024_11_26_mmfl_resnet_cifar100_with_hfl_loss', col_idx=0, color='c', line_style='-', label='MMFL simple resnet with hfl loss')
# mmfl_simple_resnet_cifar100_with_hfl_loss = read_excel('2024_11_26_mmfl_resnet_cifar100_with_hfl_vfl_loss', col_idx=0, color='r', line_style='-', label='MMFL simple resnet with hfl loss')
#
# local_simple_resnet_cifar100_multi_acc = read_excel('2024_11_24_local_resnet_cifar100_multi_test_acc', col_idx=0, color='y', line_style='-', label='local simple resnet multi loss')
# local_simple_resnet_cifar100_color_acc = read_excel('2024_11_24_local_resnet_cifar100_color_test_acc', col_idx=0, color='b', line_style='-', label='local simple resnet color acc')
# local_simple_resnet_cifar100_gray_acc = read_excel('2024_11_24_local_resnet_cifar100_gray_test_acc', col_idx=0, color='g', line_style='-', label='local simple resnet gray acc')
# # mmfl_simple_resnet_cifar100_no_client_acc = read_excel('2024_11_25_mmfl_resnet_cifar100_no_client_acc', col_idx=0, color='r', line_style='-', label='MMFL simple resnet with head acc')
# mmfl_simple_resnet_cifar100_acc = read_excel('2024_11_25_mmfl_resnet_cifar100_acc', col_idx=0, color='m', line_style='-', label='MMFL simple resnet acc')
# # mmfl_simple_resnet_cifar100_with_hfl_acc = read_excel('2024_11_26_mmfl_resnet_cifar100_with_hfl_acc', col_idx=0, color='c', line_style='-', label='MMFL simple resnet with hfl acc')
# mmfl_simple_resnet_cifar100_with_hfl_acc = read_excel('2024_11_26_mmfl_resnet_cifar100_with_hfl_vfl_acc', col_idx=0, color='r', line_style='-', label='MMFL simple resnet with hfl acc')
#
# loss_lines = [local_simple_resnet_cifar100_multi_loss, local_simple_resnet_cifar100_color_loss, local_simple_resnet_cifar100_gray_loss,
#               mmfl_simple_resnet_cifar100_loss, mmfl_simple_resnet_cifar100_with_hfl_loss]
# acc_lines = [local_simple_resnet_cifar100_multi_acc, local_simple_resnet_cifar100_color_acc, local_simple_resnet_cifar100_gray_acc,
#              mmfl_simple_resnet_cifar100_acc, mmfl_simple_resnet_cifar100_with_hfl_acc]

local_multi_loss = read_excel('2024_11_27_local_resnet_cifar100_loss', col_idx=0, color='b', line_style='--', label='local multi loss')
local_multi_acc = read_excel('2024_11_27_local_resnet_cifar100_acc', col_idx=0, color='b', line_style='--', label='local multi acc')
hfl_multi_loss = read_excel('2024_11_26_hfl_resnet_cifar100_loss', col_idx=0, color='g', line_style='--', label='hfl multi loss')
hfl_multi_acc = read_excel('2024_11_26_hfl_resnet_cifar100_acc', col_idx=0, color='g', line_style='--', label='hfl multi acc')
vfl_multi_loss = read_excel('2024_11_27_vfl_resnet_cifar100_loss', col_idx=0, color='y', line_style='--', label='vfl multi loss')
vfl_multi_acc = read_excel('2024_11_27_vfl_resnet_cifar100_acc', col_idx=0, color='y', line_style='--', label='vfl multi acc')
hfm_multi_loss = read_excel('2024_11_27_hfm_resnet_cifar100_loss', col_idx=0, color='r', line_style='--', label='hfm multi loss')
hfm_multi_acc = read_excel('2024_11_27_hfm_resnet_cifar100_acc', col_idx=0, color='r', line_style='--', label='hfm multi acc')

local_multi_loss_prox = read_excel('2024_11_27_local_resnet_cifar100_prox_loss', col_idx=0, color='b', line_style='-', label='local multi prox loss')
local_multi_acc_prox = read_excel('2024_11_27_local_resnet_cifar100_prox_acc', col_idx=0, color='b', line_style='-', label='local multi prox acc')
hfl_multi_loss_prox = read_excel('2024_11_27_hfl_resnet_cifar100_prox_loss', col_idx=0, color='g', line_style='-', label='hfl multi prox loss')
hfl_multi_acc_prox = read_excel('2024_11_27_hfl_resnet_cifar100_prox_acc', col_idx=0, color='g', line_style='-', label='hfl multi prox acc')
vfl_multi_loss_prox = read_excel('2024_11_27_vfl_resnet_cifar100_prox_loss', col_idx=0, color='y', line_style='-', label='vfl multi prox loss')
vfl_multi_acc_prox = read_excel('2024_11_27_vfl_resnet_cifar100_prox_acc', col_idx=0, color='y', line_style='-', label='vfl multi prox acc')
hfm_multi_loss_prox = read_excel('2024_11_28_hfm_resnet_cifar100_prox_loss', col_idx=0, color='r', line_style='-', label='hfm multi prox loss')
hfm_multi_acc_prox = read_excel('2024_11_28_hfm_resnet_cifar100_prox_acc', col_idx=0, color='r', line_style='-', label='hfm multi prox acc')

loss_lines = [local_multi_loss, hfl_multi_loss, vfl_multi_loss, hfm_multi_loss,
              local_multi_loss_prox, hfl_multi_loss_prox, vfl_multi_loss_prox, hfm_multi_loss_prox]
acc_lines = [local_multi_acc, hfl_multi_acc, vfl_multi_acc, hfm_multi_acc,
             local_multi_acc_prox, hfl_multi_acc_prox, vfl_multi_acc_prox, hfm_multi_acc_prox]
draw_line(loss_lines, x_label='Round', y_label='Loss', title='Loss Curve', file_name='cifar100_resnet_loss')
draw_line(acc_lines, x_label='Round', y_label='Acc', title='Accuracy Curve', file_name='cifar100_resnet_acc')
